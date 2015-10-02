# -*- coding: utf-8 -*-

from Products.Five import BrowserView
from Products.CMFCore.utils import getToolByName
from openpyxl import Workbook
from openpyxl.styles import fills, PatternFill
from openpyxl.styles import Font
from StringIO import StringIO
from DateTime import DateTime
from lx.demanda.interfaces.contents import IDemanda, IAtividade


class ExportExcelView(BrowserView):

    def __call__(self):

        self.request.response.setHeader('Content-Type', 'application/vnd.ms-excel')
        self.request.response.setHeader(
                    'Content-disposition', 'attachment;filename=Atividades.xls')

        catalog = getToolByName(self, 'portal_catalog')
        path_demandas = '/'.join(self.context.getPhysicalPath())
        ordemServico = self.request.get('ordemServico', None)
        start = self.request.get('start', None)
        end = self.request.get('end', None)

        out = StringIO()

        wb = Workbook()

        # grab the active worksheet
        ws1 = wb.active
        ws1.title = "Rotineiras"

        # Data can be assigned directly to cells
        ws1['A1'] = 'Sistema'
        ws1['B1'] = 'Ordem de Serviço'
        ws1['C1'] = 'Número da RA'
        ws1['D1'] = 'Status da RA'
        ws1['E1'] = 'Atividade'
        ws1['F1'] = 'Quantidade HST'

        # Style

        a1 = ws1['A1']
        b1 = ws1['B1']
        c1 = ws1['C1']
        d1 = ws1['D1']
        e1 = ws1['E1']
        f1 = ws1['F1']

        ft = Font(bold=True, color="FFFFFF")

        fill = PatternFill(patternType=fills.FILL_SOLID)

        a1.font = ft
        a1.fill = fill

        b1.font = ft
        b1.fill = fill

        c1.font = ft
        c1.fill = fill

        d1.font = ft
        d1.fill = fill

        e1.font = ft
        e1.fill = fill

        f1.font = ft
        f1.fill = fill

        results = catalog(object_provides=IDemanda.__identifier__,
                           path=path_demandas,
                           ordem_servico=ordemServico,
                           sort_on='chamado',
                           sort_order='reverse',)


        for i in results:
            atividade = []

            if i.Title:
                atividade.append(i.Title)
            else:
                atividade.append('')

            if i.ordem_servico:
                atividade.append(i.ordem_servico)
            else:
                atividade.append('')

            if i.chamado:
                atividade.append(i.chamado)
            else:
                atividade.append('')

            if i.status_ra:
                atividade.append(i.status_ra)
            else:
                atividade.append('')

            if i.atividade:
                atividade.append(i.atividade)
            else:
                atividade.append('')

            if i.quantHST:
                atividade.append(i.quantHST)
            else:
                atividade.append('')

            ws1.append(atividade)

        # Rows can also be appended
        # ws1.append([1, 2, 3])
        # ws1.append([4, 5, 6])

        if start and end:
            ws2 = wb.create_sheet(title='Projetizadas')

            ws2['A1'] = 'Atividade'
            ws2['B1'] = 'Projeto'
            ws2['C1'] = 'Data início'
            ws2['D1'] = 'Data fim'
            ws2['E1'] = 'Duração'
            ws2['F1'] = 'Quantidade HST'

            # Style

            a1 = ws2['A1']
            b1 = ws2['B1']
            c1 = ws2['C1']
            d1 = ws2['D1']
            e1 = ws2['E1']
            f1 = ws2['F1']

            ft = Font(bold=True, color="FFFFFF")

            fill = PatternFill(patternType=fills.FILL_SOLID)

            a1.font = ft
            a1.fill = fill

            b1.font = ft
            b1.fill = fill

            c1.font = ft
            c1.fill = fill

            d1.font = ft
            d1.fill = fill

            e1.font = ft
            e1.fill = fill

            f1.font = ft
            f1.fill = fill

            first_date = DateTime(start, datefmt='international')
            last_date = DateTime(end + ' 23:59:59', datefmt='international')
            results = catalog(object_provides=IAtividade.__identifier__,
                               path=path_demandas,
                               data_inicio={'query': first_date, 'range': 'min'},
                               data_fim={'query': last_date, 'range': 'max'},
                               sort_on='data_inicio')

            for i in results:
                atividade = []

                if i.Title:
                    atividade.append(i.Title)
                else:
                    atividade.append('')

                if i.projeto:
                    atividade.append(i.projeto)
                else:
                    atividade.append('')

                if i.data_inicio:
                    data_inicio = i.data_inicio.strftime('%d/%m/%Y %H:%M')
                    atividade.append(data_inicio)
                else:
                    atividade.append('')

                if i.data_fim:
                    data_fim = i.data_fim.strftime('%d/%m/%Y %H:%M')
                    atividade.append(data_fim)
                else:
                    atividade.append('')

                if i.duracao:
                    atividade.append(i.duracao)
                else:
                    atividade.append('')

                if i.quantHST:
                    atividade.append(i.quantHST)
                else:
                    atividade.append('')

                ws2.append(atividade)

        # Save the file
        wb.save(out)

        return out.getvalue()